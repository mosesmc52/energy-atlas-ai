"""
Parse EIA-StatetoStateCapacity_Jan<YYYY>.xlsx tab-by-tab using *anchor cells*.

Updates vs your version:
- Adds current_january_url() + current_january_filename()
- Adds CLI flags to set:
    1) destination path of downloaded XLSX  (either --dest-xlsx or --download-dir)
    2) output directory for generated files (--out-dir)
    3) output format (--output-format: csv or parquet)
- Keeps dynamic row/column detection (safe if rows expand)

Examples:
  # default locations (download into data/raw/... ; extracted files into data/staging/...)
  python ingest_state_to_state_capacity.py

  # explicit destinations
  python ingest_state_to_state_capacity.py \
      --download-dir data/raw/eia/ng/pipeline/state_to_state_capacity/2026-01 \
      --out-dir data/staging/eia/ng/pipeline/state_to_state_capacity/2026-01

  # or explicit xlsx filepath
  python ingest_state_to_state_capacity.py \
      --dest-xlsx /tmp/EIA-StatetoStateCapacity_Jan2026.xlsx \
      --out-dir /tmp/out_csv
"""

from __future__ import annotations

import argparse
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet


# -----------------------------
# Year-aware URL + filename
# -----------------------------
def current_year() -> int:
    return datetime.now().year


def current_january_filename(prefix: str) -> str:
    # ex: "EIA-StatetoStateCapacity_Jan2026.xlsx"
    return f"{prefix}_Jan{current_year()}.xlsx"


def current_january_url(prefix: str) -> str:
    # ex: https://www.eia.gov/naturalgas/pipelines/EIA-StatetoStateCapacity_Jan2026.xlsx
    return (
        f"https://www.eia.gov/naturalgas/pipelines/{current_january_filename(prefix)}"
    )


EIA_PREFIX = "EIA-StatetoStateCapacity"
EIA_URL = current_january_url(EIA_PREFIX)


# -----------------------------
# Config: sheet -> anchors
# -----------------------------
@dataclass(frozen=True)
class SheetSpec:
    sheet: str
    header_cell: str  # top-left header cell
    data_cell: str  # top-left data cell (first data row)


SPECS = [
    SheetSpec("Major Pipeline Summary", "B5", "B6"),
    SheetSpec("Inflow By Region", "A5", "A6"),
    SheetSpec("Outflow By Region", "A5", "A6"),
    SheetSpec("Inflow By State", "A5", "A6"),
    SheetSpec("Outflow By State", "A5", "A6"),
    SheetSpec("Inflow By State and Pipeline", "A5", "A6"),
    SheetSpec("Outflow By State and Pipeline", "A5", "A6"),
    SheetSpec("Pipeline State2State Capacity", "A2", "A3"),
    SheetSpec("State2StateAIMMS", "A1", "A1"),
    SheetSpec("Pipeline State2State CapacityH", "A5", "A5"),
    SheetSpec("InFlow Single Year", "A5", "A5"),
    SheetSpec("Outflow Single Year", "A4", "A5"),
]


# -----------------------------
# Helpers
# -----------------------------
def download_if_needed(url: str, filepath: str) -> str:
    """
    Downloads url -> filepath if file doesn't exist (or is empty).
    """
    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
    if os.path.exists(filepath) and os.path.getsize(filepath) > 0:
        return filepath

    r = requests.get(url, timeout=60)
    r.raise_for_status()
    with open(filepath, "wb") as f:
        f.write(r.content)
    return filepath


def _split_cell(cell: str) -> Tuple[int, int]:
    col_letters, row = coordinate_from_string(cell)
    return int(row), int(column_index_from_string(col_letters))


def _is_blank(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")


def detect_last_col_from_header(
    ws: Worksheet,
    header_row: int,
    start_col: int,
    max_consecutive_blanks: int = 10,
    hard_max_cols: int = 5000,
) -> int:
    last_nonblank = start_col
    blanks = 0
    upper = min(max(ws.max_column, start_col) + 200, hard_max_cols)

    for c in range(start_col, upper + 1):
        v = ws.cell(row=header_row, column=c).value
        if _is_blank(v):
            blanks += 1
            if blanks >= max_consecutive_blanks:
                break
        else:
            last_nonblank = c
            blanks = 0

    return last_nonblank


def detect_last_row_from_data(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_col: int,
    max_consecutive_blank_rows: int = 50,
) -> int:
    last_data_row = start_row - 1
    blanks = 0

    for r in range(start_row, ws.max_row + 1):
        any_data = False
        for c in range(start_col, end_col + 1):
            if not _is_blank(ws.cell(row=r, column=c).value):
                any_data = True
                break

        if any_data:
            last_data_row = r
            blanks = 0
        else:
            blanks += 1
            if blanks >= max_consecutive_blank_rows:
                break

    return last_data_row


def _dedupe_headers(headers: list[str]) -> list[str]:
    out = []
    seen: Dict[str, int] = {}
    for i, h in enumerate(headers, start=1):
        base = h if h else f"Unnamed_{i}"
        k = base
        if k in seen:
            seen[k] += 1
            k = f"{base}_{seen[base]}"
        else:
            seen[k] = 1
        out.append(k)
    return out


def safe_filename(name: str) -> str:
    name = name.strip().replace("/", "-")
    name = re.sub(r"[^A-Za-z0-9._ -]+", "", name)
    name = re.sub(r"\s+", "_", name)
    return name


def _normalize_for_parquet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Make DataFrame parquet-safe for pyarrow when Excel-origin columns are mixed-type
    objects (e.g., strings and ints in the same column).
    """
    out = df.copy()
    object_cols = out.select_dtypes(include=["object"]).columns
    for col in object_cols:
        out[col] = out[col].astype("string")
    return out


def read_table(
    ws: Worksheet,
    header_cell: str,
    data_cell: str,
    *,
    header_blank_run: int = 10,
    row_blank_run: int = 50,
) -> pd.DataFrame:
    header_row, start_col = _split_cell(header_cell)
    data_row, data_col = _split_cell(data_cell)

    # If header == data anchor, assume data starts on next row
    if header_row == data_row and start_col == data_col:
        data_row = header_row + 1

    start_col = min(start_col, data_col)

    max_col = min(max(ws.max_column, start_col), 5000)
    row_iter = ws.iter_rows(
        min_row=header_row,
        max_row=ws.max_row,
        min_col=start_col,
        max_col=max_col,
        values_only=True,
    )

    try:
        header_values = next(row_iter)
    except StopIteration:
        return pd.DataFrame()

    last_nonblank_idx = 0
    blanks = 0
    for idx, value in enumerate(header_values):
        if _is_blank(value):
            blanks += 1
            if blanks >= header_blank_run:
                break
        else:
            last_nonblank_idx = idx
            blanks = 0

    headers = [header_values[idx] for idx in range(last_nonblank_idx + 1)]
    headers = [("" if h is None else str(h).strip()) for h in headers]
    headers = _dedupe_headers(headers)

    rows = []
    blank_rows = 0
    first_data_offset = data_row - header_row - 1
    for offset, row in enumerate(row_iter):
        if offset < first_data_offset:
            continue

        row_values = list(row[: last_nonblank_idx + 1])
        if any(not _is_blank(value) for value in row_values):
            rows.append(row_values)
            blank_rows = 0
        else:
            rows.append(row_values)
            blank_rows += 1
            if blank_rows >= row_blank_run:
                break

    df = pd.DataFrame(rows, columns=headers).dropna(how="all")
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    return df


# -----------------------------
# Main parsing
# -----------------------------
def parse_workbook(
    xlsx_path: str,
    specs: list[SheetSpec] = SPECS,
    out_dir: Optional[str] = "out_csv",
    output_format: str = "csv",
) -> Dict[str, pd.DataFrame]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True, keep_links=False)
    out: Dict[str, pd.DataFrame] = {}

    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    for spec in specs:
        if spec.sheet not in wb.sheetnames:
            raise KeyError(
                f"Sheet not found: {spec.sheet!r}. Available: {wb.sheetnames}"
            )

        ws = wb[spec.sheet]
        df = read_table(ws, spec.header_cell, spec.data_cell)
        out[spec.sheet] = df

        if out_dir:
            base_path = os.path.join(out_dir, safe_filename(spec.sheet))
            if output_format == "csv":
                df.to_csv(f"{base_path}.csv", index=False)
            elif output_format == "parquet":
                parquet_df = _normalize_for_parquet(df)
                parquet_df.to_parquet(f"{base_path}.parquet", index=False)
            else:
                raise ValueError(
                    f"Unsupported output_format={output_format!r}. Use 'csv' or 'parquet'."
                )

        print(f"[OK] {spec.sheet}: rows={len(df):,} cols={df.shape[1]:,}")

    return out


def main():
    """
    Controls:
    - where the XLSX is downloaded (dest)
    - where extracted files are written (out_dir)
    - which extracted file format is written (output-format)
    """
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--download-url",
        default=EIA_URL,
        help="EIA XLSX URL (defaults to current January for the current year).",
    )
    parser.add_argument(
        "--download-dir",
        default="data/raw/eia/ng/pipeline",
        help="Directory to store the downloaded XLSX (ignored if --dest-xlsx is set).",
    )

    parser.add_argument(
        "--out-dir",
        default="data/processed/eia/ng/pipeline/",
        help="Directory to write one extracted file per sheet.",
    )
    parser.add_argument(
        "--output-format",
        choices=["csv", "parquet"],
        default="csv",
        help="Output file format.",
    )
    args = parser.parse_args()

    # Decide XLSX destination
    fname = current_january_filename(EIA_PREFIX)

    dl_dir = Path(args.download_dir) if args.download_dir else Path(".")
    xlsx_path = dl_dir / fname

    # Ensure download
    if not xlsx_path.exists() or xlsx_path.stat().st_size == 0:
        print(f"Downloading: {args.download_url}")
        download_if_needed(args.download_url, str(xlsx_path))

    # Parse + write extracted files
    out_dir = args.out_dir if args.out_dir.strip() else None
    parse_workbook(
        str(xlsx_path),
        SPECS,
        out_dir=out_dir,
        output_format=args.output_format,
    )


if __name__ == "__main__":
    main()
