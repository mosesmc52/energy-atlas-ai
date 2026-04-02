"""
EIA Natural Gas Pipeline Projects (Jan 2026) parser

Sheets:
- "Natural Gas Pipeline Projects"  (Header A2, Data starts A3)
- "Historical Projects (1996-2024)" (Header A2, Data starts A3)

Row counts may expand -> we detect:
- last column from header row (scan right until a run of blank header cells)
- last row from data region (scan down until a run of blank rows)

Usage:
  python scripts/eia/ng/pipeline/ingest_pipeline_projects.py \
      --xlsx /path/to/EIA-NaturalGasPipelineProjects_Jan2026.xlsx \
      --out out_csv

Or (in this environment):
  --xlsx /mnt/data/EIA-NaturalGasPipelineProjects_Jan2026.xlsx

Requires:
  pip install pandas openpyxl pyarrow
"""

from __future__ import annotations

import argparse
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, Tuple

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet


# -----------------------------
# Config
# -----------------------------
@dataclass(frozen=True)
class SheetSpec:
    sheet: str
    header_cell: str
    data_cell: str  # top-left of first data row


SPECS = [
    SheetSpec("Natural Gas Pipeline Projects", "A2", "A3"),
    SheetSpec("Historical Projects (1996-2024)", "A2", "A3"),
]


# -----------------------------
# Helpers
# -----------------------------


def current_year() -> int:
    return datetime.now().year


def current_january_filename(prefix: str) -> str:
    return f"{prefix}_Jan{current_year()}.xlsx"


def current_january_url(prefix: str) -> str:
    return (
        f"https://www.eia.gov/naturalgas/pipelines/{current_january_filename(prefix)}"
    )


def download_if_needed(url: str, dest: Path) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)

    if dest.exists() and dest.stat().st_size > 0:
        return dest

    print(f"[download] {url}")
    r = requests.get(url, timeout=60)
    r.raise_for_status()

    with open(dest, "wb") as f:
        f.write(r.content)

    return dest


def _split_cell(cell: str) -> Tuple[int, int]:
    col_letters, row = coordinate_from_string(cell)
    return int(row), int(column_index_from_string(col_letters))


def _is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    return False


def detect_last_col_from_header(
    ws: Worksheet,
    header_row: int,
    start_col: int,
    max_consecutive_blanks: int = 12,
    hard_max_cols: int = 5000,
) -> int:
    """
    Scan right from header_row/start_col until we hit N consecutive blank header cells.
    Returns last non-blank header column index.
    """
    last_nonblank = start_col
    blanks = 0
    upper = min(max(ws.max_column, start_col) + 200, hard_max_cols)

    for c in range(start_col, upper + 1):
        v = ws.cell(row=header_row, column=c).value
        if _is_blank(v):
            blanks += 1
            if blanks >= max_consecutive_blanks:
                break
        else:
            last_nonblank = c
            blanks = 0
    return last_nonblank


def detect_last_row_from_data(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    end_col: int,
    max_consecutive_blank_rows: int = 80,
) -> int:
    """
    Scan downward; a row is "data" if any cell in [start_col..end_col] is non-blank.
    Stop after N consecutive blank rows.
    """
    last_data_row = start_row - 1
    blanks = 0

    for r in range(start_row, ws.max_row + 1):
        any_data = False
        for c in range(start_col, end_col + 1):
            if not _is_blank(ws.cell(row=r, column=c).value):
                any_data = True
                break

        if any_data:
            last_data_row = r
            blanks = 0
        else:
            blanks += 1
            if blanks >= max_consecutive_blank_rows:
                break

    return last_data_row


def _dedupe_headers(headers: list[str]) -> list[str]:
    """
    Make headers unique: ["State","State",""] -> ["State","State_2","Unnamed_3"]
    """
    out: list[str] = []
    seen: Dict[str, int] = {}
    for i, h in enumerate(headers, start=1):
        base = (h or "").strip() or f"Unnamed_{i}"
        if base in seen:
            seen[base] += 1
            out.append(f"{base}_{seen[base]}")
        else:
            seen[base] = 1
            out.append(base)
    return out


def safe_filename(name: str) -> str:
    name = name.strip().replace("/", "-")
    name = re.sub(r"[^A-Za-z0-9._ -]+", "", name)
    name = re.sub(r"\s+", "_", name)
    return name


def read_table(ws: Worksheet, header_cell: str, data_cell: str) -> pd.DataFrame:
    header_row, start_col = _split_cell(header_cell)
    data_row, data_col = _split_cell(data_cell)
    start_col = min(start_col, data_col)

    max_col = min(max(ws.max_column, start_col), 5000)
    row_iter = ws.iter_rows(
        min_row=header_row,
        max_row=ws.max_row,
        min_col=start_col,
        max_col=max_col,
        values_only=True,
    )

    try:
        header_values = next(row_iter)
    except StopIteration:
        return pd.DataFrame()

    last_nonblank_idx = 0
    blanks = 0
    for idx, value in enumerate(header_values):
        if _is_blank(value):
            blanks += 1
            if blanks >= 12:
                break
        else:
            last_nonblank_idx = idx
            blanks = 0

    headers = [header_values[idx] for idx in range(last_nonblank_idx + 1)]
    headers = [("" if h is None else str(h).strip()) for h in headers]
    headers = _dedupe_headers(headers)

    rows = []
    blank_rows = 0
    first_data_offset = data_row - header_row - 1
    for offset, row in enumerate(row_iter):
        if offset < first_data_offset:
            continue

        row_values = list(row[: last_nonblank_idx + 1])
        if any(not _is_blank(value) for value in row_values):
            rows.append(row_values)
            blank_rows = 0
        else:
            rows.append(row_values)
            blank_rows += 1
            if blank_rows >= 80:
                break

    df = pd.DataFrame(rows, columns=headers)
    df = df.dropna(how="all")
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    return df


def _normalize_object_cols_for_parquet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Make object columns Arrow-safe:
    - Normalize common NA-like tokens to missing values
    - Convert fully numeric-like object columns to numeric dtype
    - Convert remaining object columns to pandas string dtype
    """
    out = df.copy()
    na_tokens = {"", "na", "n/a", "none", "null", "nan", "n.a."}

    object_cols = out.select_dtypes(include=["object"]).columns.tolist()
    for col in object_cols:
        s = out[col]

        # normalize string noise and NA markers
        s_norm = s.map(
            lambda v: (
                pd.NA
                if isinstance(v, str) and v.strip().lower() in na_tokens
                else (v.strip() if isinstance(v, str) else v)
            )
        )

        # if all non-null values are numeric-like, coerce to numeric
        non_null = s_norm.dropna()
        if len(non_null) > 0:
            as_num = pd.to_numeric(non_null, errors="coerce")
            if as_num.notna().all():
                out[col] = pd.to_numeric(s_norm, errors="coerce")
                continue

        # otherwise keep as string for stable parquet typing
        out[col] = s_norm.astype("string")

    return out


# -----------------------------
# Main
# -----------------------------
def parse_workbook(
    xlsx_path: str,
    out_dir: Optional[str] = "out_csv",
    output_format: str = "csv",
) -> Dict[str, pd.DataFrame]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True, keep_links=False)

    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    out: Dict[str, pd.DataFrame] = {}
    for spec in SPECS:
        if spec.sheet not in wb.sheetnames:
            raise KeyError(
                f"Sheet not found: {spec.sheet!r}. Available: {wb.sheetnames}"
            )

        ws = wb[spec.sheet]
        df = read_table(ws, spec.header_cell, spec.data_cell)
        out[spec.sheet] = df

        if out_dir:
            base_path = os.path.join(out_dir, safe_filename(spec.sheet))
            if output_format == "csv":
                df.to_csv(f"{base_path}.csv", index=False)
            elif output_format == "parquet":
                df_parquet = _normalize_object_cols_for_parquet(df)
                df_parquet.to_parquet(f"{base_path}.parquet", index=False)
            else:
                raise ValueError(
                    f"Unsupported output_format={output_format!r}. Use 'csv' or 'parquet'."
                )

        print(f"[OK] {spec.sheet}: rows={len(df):,} cols={df.shape[1]:,}")

    return out


def main():
    p = argparse.ArgumentParser()
    p.add_argument(
        "--out",
        default="data/processed/eia/ng/pipeline/",
        help="Output folder for extracted files",
    )
    p.add_argument(
        "--output-format",
        choices=["csv", "parquet"],
        default="parquet",
        help="Output file format",
    )
    p.add_argument("--raw-dir", default="data/raw/eia/ng/pipeline")
    args = p.parse_args()

    prefix = "EIA-NaturalGasPipelineProjects"

    url = current_january_url(prefix)
    fname = current_january_filename(prefix)

    raw_dir = Path(args.raw_dir)
    xlsx_path = raw_dir / fname

    download_if_needed(url, xlsx_path)

    parse_workbook(str(xlsx_path), out_dir=args.out, output_format=args.output_format)


if __name__ == "__main__":
    main()
